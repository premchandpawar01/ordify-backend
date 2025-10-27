# admin_dashboard.py
# v5.9.2 - Fixed 'bills' vs 'monthly_bills' page_key mismatch
#        - Corrected stylesheet for CardFrame widgets
#        - Removed Client search bar connection per user request
# UPDATED: Changed default filter to "All Time"
# UPDATED: Added style for transparent Add buttons
# UPDATED: Implemented product image view dialog and new search bar styles
# (FIXED: Product image not scaling correctly on dialog open)
# UPDATED: Added new stylesheet rules for BaseDialog
# UPDATED: Refactored ProductDetailDialog into its own file and based it on BaseDialog
# UPDATED: Removed unsupported 'box-shadow' property

import sys
import requests
from datetime import datetime, timedelta
import webbrowser
import os
import tempfile
import csv
import openpyxl
from openpyxl.utils import get_column_letter
import threading # Import threading for image download

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QFrame, QPushButton, QTableWidget, QTableWidgetItem,
    QMessageBox, QDialogButtonBox, QStackedWidget, QLabel,
    QFileDialog, QDateEdit, QCheckBox, QFormLayout, QDialog,
    QProgressBar
)
from PyQt6.QtCore import Qt, QDate, QSize, QPropertyAnimation, QEasingCurve, pyqtSignal
from PyQt6.QtGui import QColor, QIcon, QFont, QCursor, QPixmap

# --- Import Page UIs ---
from pages.dashboard_page import DashboardPage
from pages.clients_page import ClientsPage
from pages.products_page import ProductsPage
from pages.orders_page import OrdersPage
from pages.challans_page import ChallansPage
from pages.monthly_bills_page import MonthlyBillsPage

# --- Import Dialog UIs ---
from dialogs.product_dialog import ProductDialog
from dialogs.client_dialog import ClientDialog
from dialogs.client_pricing_dialog import ClientPricingDialog
from dialogs.mark_as_paid_dialog import MarkAsPaidDialog
from dialogs.filter_dialog import FilterDialog
from dialogs.base_dialog import BaseDialog 
# NEW: Import the refactored Product Detail Dialog
from dialogs.product_detail_dialog import ProductDetailDialog
# NEW: Import the Order Detail Dialog
from dialogs.order_detail_dialog import OrderDetailDialog

# ---
# --- ICON DEFINITIONS ---
# ---
ICON_PATH = "icons/"
ICON_DASHBOARD = os.path.join(ICON_PATH, "dashboard.png")
ICON_CLIENTS = os.path.join(ICON_PATH, "clients.png")
ICON_PRODUCTS = os.path.join(ICON_PATH, "products.png")
ICON_ORDERS = os.path.join(ICON_PATH, "orders.png")
ICON_CHALLANS = os.path.join(ICON_PATH, "challans.png")
ICON_BILLS = os.path.join(ICON_PATH, "bills.png")
ICON_COLLAPSE = os.path.join(ICON_PATH, "collapse.png")
ICON_EXPAND = os.path.join(ICON_PATH, "expand.png")
ICON_VIEW = os.path.join(ICON_PATH, "view.png")
ICON_EDIT = os.path.join(ICON_PATH, "edit.png")
ICON_DELETE = os.path.join(ICON_PATH, "delete.png")
ICON_PRICING = os.path.join(ICON_PATH, "pricing.png")
ICON_PDF = os.path.join(ICON_PATH, "pdf.png")
ICON_RESET = os.path.join(ICON_PATH, "reset.png")
ICON_PAID = os.path.join(ICON_PATH, "paid.png")
ICON_FUNNEL = os.path.join(ICON_PATH, "filter.png")
ICON_CSV = os.path.join(ICON_PATH, "csv.png")
ICON_EXCEL = os.path.join(ICON_PATH, "excel.png")
# NEW: Add an icon for the "Add" button
ICON_ADD = os.path.join(ICON_PATH, "add.png") # You'll need to find an "add.png" icon

API_BASE_URL = "https://ordify-api.onrender.com"

# ===================================================================
# --- REMOVED ProductDetailDialog class ---
# It is now imported from dialogs/product_detail_dialog.py
# ===================================================================


# --- Main Application Window ---
class AdminDashboard(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Admin Dashboard - OrdIfY")
        self.setGeometry(100, 100, 1400, 800)

        # --- Add icon paths as class attributes for pages to access ---
        self.ICON_DASHBOARD = ICON_DASHBOARD
        self.ICON_CLIENTS = ICON_CLIENTS
        self.ICON_PRODUCTS = ICON_PRODUCTS
        self.ICON_ORDERS = ICON_ORDERS
        self.ICON_BILLS = ICON_BILLS
        self.ICON_EDIT = ICON_EDIT
        self.ICON_DELETE = ICON_DELETE
        self.ICON_PRICING = ICON_PRICING
        self.ICON_VIEW = ICON_VIEW
        self.ICON_CHALLANS = ICON_CHALLANS
        self.ICON_PDF = ICON_PDF
        self.ICON_RESET = ICON_RESET
        self.ICON_PAID = ICON_PAID
        self.ICON_FUNNEL = ICON_FUNNEL
        self.ICON_CSV = ICON_CSV
        self.ICON_EXCEL = ICON_EXCEL
        self.ICON_ADD = ICON_ADD # Add new icon
        self.API_BASE_URL = API_BASE_URL

        # --- Main Layout ---
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QHBoxLayout(self.central_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)

        self.clients_list = []
        self.products_list = []
        self.sidebar_expanded = True

        self._current_page = {
            'orders': 1,
            'challans': 1,
            'monthly_bills': 1,
        }
        self._total_pages = {
            'orders': 1,
            'challans': 1,
            'monthly_bills': 1,
        }

        # NEW: Store current filter settings
        self._filter_settings = {
            'orders': {'type': 'All Time', 'start': None, 'end': None},
            'challans': {'type': 'All Time', 'start': None, 'end': None},
            'monthly_bills': {'type': 'All Time', 'start': None, 'end': None},
        }

        # --- 1. Create Sidebar ---
        self.create_sidebar()
        self.main_layout.addWidget(self.sidebar_frame)

        # --- 2. Create Main Content Area ---
        self.main_content_frame = QFrame()
        self.main_content_frame.setObjectName("ContentFrame")
        self.main_content_layout = QVBoxLayout(self.main_content_frame)

        self.stacked_widget = QStackedWidget()
        self.main_content_layout.addWidget(self.stacked_widget)

        # --- Instantiate Page Widgets ---
        self.dashboard_page = DashboardPage(self)
        self.clients_page = ClientsPage(self)
        self.products_page = ProductsPage(self)
        self.orders_page = OrdersPage(self)
        self.challans_page = ChallansPage(self)
        self.monthly_bills_page = MonthlyBillsPage(self)

        # --- Add Page Widgets to Stack ---
        self.stacked_widget.addWidget(self.dashboard_page)
        self.stacked_widget.addWidget(self.clients_page)
        self.stacked_widget.addWidget(self.products_page)
        self.stacked_widget.addWidget(self.orders_page)
        self.stacked_widget.addWidget(self.challans_page)
        self.stacked_widget.addWidget(self.monthly_bills_page)

        self.main_layout.addWidget(self.main_content_frame, 1)

        # --- 3. Connect Sidebar Buttons ---
        self.connect_sidebar_buttons()

        # --- 4. Apply Styles ---
        self.apply_styles()

        # --- 5. Initial Data Load ---
        self.refresh_all_data()
    # <<< End of __init__ method

    # --- Helper to fetch data ---
    def fetch_generic_details(self, endpoint, params=None):
        """Fetches data from the API and handles errors."""
        try:
            response = requests.get(f"{API_BASE_URL}{endpoint}", params=params)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            error_message = f"API Error: Could not fetch details from {endpoint}.\n\nReason: {e}"
            if e.response is not None:
                try:
                    api_error = e.response.json().get('error', e.response.text)
                    error_message += f"\n\nAPI Response: {api_error}"
                except requests.exceptions.JSONDecodeError:
                    error_message += f"\n\nAPI Response: (Non-JSON) {e.response.text}"
            QMessageBox.warning(self, "API Connection Error", error_message)
            return None

    def create_sidebar(self):
        self.sidebar_frame = QFrame()
        self.sidebar_frame.setObjectName("SidebarFrame")
        self.sidebar_frame.setFixedWidth(220)

        self.sidebar_layout = QVBoxLayout(self.sidebar_frame)
        self.sidebar_layout.setContentsMargins(10, 10, 10, 10)
        self.sidebar_layout.setSpacing(10)

        self.brand_label = QLabel("OrdIfY")
        self.brand_label.setObjectName("BrandLabel")
        self.brand_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.sidebar_layout.addWidget(self.brand_label)

        self.nav_buttons = []
        self.btn_dashboard = QPushButton("Dashboard")
        self.btn_clients = QPushButton("Clients")
        self.btn_products = QPushButton("Products")
        self.btn_orders = QPushButton("Orders")
        self.btn_challans = QPushButton("Challans")
        self.btn_bills = QPushButton("Monthly Bills")

        self.nav_buttons.append((self.btn_dashboard, "Dashboard", ICON_DASHBOARD))
        self.nav_buttons.append((self.btn_clients, "Clients", ICON_CLIENTS))
        self.nav_buttons.append((self.btn_products, "Products", ICON_PRODUCTS))
        self.nav_buttons.append((self.btn_orders, "Orders", ICON_ORDERS))
        self.nav_buttons.append((self.btn_challans, "Challans", ICON_CHALLANS))
        self.nav_buttons.append((self.btn_bills, "Monthly Bills", ICON_BILLS))

        icon_size = QSize(20, 20)

        for btn, text, icon_path in self.nav_buttons:
            btn.setObjectName("SidebarButton")
            btn.setIcon(QIcon(icon_path))
            btn.setIconSize(icon_size)
            btn.setCheckable(True)
            btn.setAutoExclusive(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            self.sidebar_layout.addWidget(btn)

        self.sidebar_layout.addStretch()

        self.toggle_btn = QPushButton("Collapse")
        self.toggle_btn.setObjectName("SidebarButton")
        self.toggle_btn.setIcon(QIcon(ICON_COLLAPSE))
        self.toggle_btn.setIconSize(icon_size)
        self.toggle_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.toggle_btn.clicked.connect(self.toggle_sidebar)
        self.sidebar_layout.addWidget(self.toggle_btn)

        self.btn_dashboard.setChecked(True)

    def connect_sidebar_buttons(self):
        # Connect buttons to refresh data AND switch pages
        self.btn_dashboard.clicked.connect(lambda: (
            self.refresh_dashboard_data(),
            self.stacked_widget.setCurrentWidget(self.dashboard_page)
        ))
        self.btn_clients.clicked.connect(lambda: (
            self.refresh_clients_data(),
            self.stacked_widget.setCurrentWidget(self.clients_page)
        ))
        self.btn_products.clicked.connect(lambda: (
            self.refresh_products_data(),
            self.stacked_widget.setCurrentWidget(self.products_page)
        ))

        self.btn_orders.clicked.connect(lambda: (
            self.refresh_orders_data(page_num=1), # Reset to page 1
            self.stacked_widget.setCurrentWidget(self.orders_page)
        ))
        self.btn_challans.clicked.connect(lambda: (
            self.refresh_challans_data(page_num=1), # Reset to page 1
            self.stacked_widget.setCurrentWidget(self.challans_page)
        ))
        self.btn_bills.clicked.connect(lambda: (
            self.refresh_monthly_bills_data(page_num=1), # Reset to page 1
            self.stacked_widget.setCurrentWidget(self.monthly_bills_page)
        ))

        # --- Connect Filter/Pagination UI ---

        # --- Orders Page ---
        try:
            # Connect the new filter dialog workflow
            self.orders_page.filter_button.clicked.connect(lambda: self.open_filter_dialog('orders'))
            self.orders_page.reset_button.clicked.connect(lambda: self.reset_page_filter('orders'))

            self.orders_page.next_button.clicked.connect(lambda: self.go_to_next_page('orders'))
            self.orders_page.prev_button.clicked.connect(lambda: self.go_to_prev_page('orders'))

            self.orders_page.export_csv_button.clicked.connect(
                lambda: self.export_table_to_csv(self.orders_page.table, "Orders_Export.csv")
            )
            self.orders_page.export_excel_button.clicked.connect(
                lambda: self.export_table_to_xlsx(self.orders_page.table, "Orders_Export.xlsx")
            )
        except AttributeError as e:
            print(f"Warning: Could not connect Orders page UI elements. {e}")

        # --- Challans Page ---
        try:
            self.challans_page.filter_button.clicked.connect(lambda: self.open_filter_dialog('challans'))
            self.challans_page.reset_button.clicked.connect(lambda: self.reset_page_filter('challans'))

            self.challans_page.next_button.clicked.connect(lambda: self.go_to_next_page('challans'))
            self.challans_page.prev_button.clicked.connect(lambda: self.go_to_prev_page('challans'))

            self.challans_page.export_csv_button.clicked.connect(
                lambda: self.export_table_to_csv(self.challans_page.table, "Challans_Export.csv")
            )
            self.challans_page.export_excel_button.clicked.connect(
                lambda: self.export_table_to_xlsx(self.challans_page.table, "Challans_Export.xlsx")
            )
        except AttributeError as e:
            print(f"Warning: Could not connect Challans page UI elements. {e}")

        # --- Monthly Bills Page ---
        try:
            self.monthly_bills_page.filter_button.clicked.connect(lambda: self.open_filter_dialog('monthly_bills'))
            self.monthly_bills_page.reset_button.clicked.connect(lambda: self.reset_page_filter('monthly_bills'))

            self.monthly_bills_page.next_button.clicked.connect(lambda: self.go_to_next_page('monthly_bills'))
            self.monthly_bills_page.prev_button.clicked.connect(lambda: self.go_to_prev_page('monthly_bills'))

            self.monthly_bills_page.export_csv_button.clicked.connect(
                lambda: self.export_table_to_csv(self.monthly_bills_page.table, "Bills_Export.csv")
            )
            self.monthly_bills_page.export_excel_button.clicked.connect(
                lambda: self.export_table_to_xlsx(self.monthly_bills_page.table, "Bills_Export.xlsx")
            )
        except AttributeError as e:
            print(f"Warning: Could not connect Bills page UI elements. {e}")

        # --- Connect other page export buttons (Clients, Products) ---
        try:
            self.clients_page.export_csv_button.clicked.connect(
                lambda: self.export_table_to_csv(self.clients_page.table, "Clients_Export.csv")
            )
            self.clients_page.export_excel_button.clicked.connect(
                lambda: self.export_table_to_xlsx(self.clients_page.table, "Clients_Export.xlsx")
            )
            # No search bar to connect, per user request
        except AttributeError as e:
            print(f"Warning: Could not connect Clients page export buttons. {e}")

        try:
            self.products_page.export_csv_button.clicked.connect(
                lambda: self.export_table_to_csv(self.products_page.table, "Products_Export.csv")
            )
            self.products_page.export_excel_button.clicked.connect(
                lambda: self.export_table_to_xlsx(self.products_page.table, "Products_Export.xlsx")
            )
            # Search bar connection is handled *inside* products_page.py
        except AttributeError as e:
            print(f"Warning: Could not connect Products page export buttons. {e}")

    def toggle_sidebar(self):
        icon_size = QSize(20, 20)
        if self.sidebar_expanded:
            new_width = 60
            icon_path = ICON_EXPAND
            toggle_tooltip = "Expand"
            brand_text = ""
            button_text_map = {btn: "" for btn, _, _ in self.nav_buttons}
            toggle_text = ""
        else:
            new_width = 220
            icon_path = ICON_COLLAPSE
            toggle_tooltip = "Collapse"
            brand_text = "OrdIfY"
            button_text_map = {btn: text for btn, text, _ in self.nav_buttons}
            toggle_text = "Collapse"

        self.min_anim = QPropertyAnimation(self.sidebar_frame, b"minimumWidth")
        self.min_anim.setDuration(300)
        self.min_anim.setStartValue(self.sidebar_frame.width())
        self.min_anim.setEndValue(new_width)
        self.min_anim.setEasingCurve(QEasingCurve.Type.InOutCubic)

        self.max_anim = QPropertyAnimation(self.sidebar_frame, b"maximumWidth")
        self.max_anim.setDuration(300)
        self.max_anim.setStartValue(self.sidebar_frame.width())
        self.max_anim.setEndValue(new_width)
        self.max_anim.setEasingCurve(QEasingCurve.Type.InOutCubic)

        self.min_anim.start()
        self.max_anim.start()

        self.toggle_btn.setText(toggle_text)
        self.toggle_btn.setIcon(QIcon(icon_path))
        self.toggle_btn.setIconSize(icon_size)
        self.toggle_btn.setToolTip(toggle_tooltip)

        self.brand_label.setText(brand_text)
        self.brand_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        for btn, text, icon_path in self.nav_buttons:
            btn.setText(button_text_map[btn])
            btn.setToolTip(text if new_width == 60 else "")

        self.sidebar_expanded = not self.sidebar_expanded

    # --- Data Refresh Methods ---
    def clear_layout(self, layout):
         if layout is not None:
            if hasattr(layout, 'count'):
                while layout.count():
                    item = layout.takeAt(0)
                    widget = item.widget()
                    if widget is not None:
                        widget.deleteLater()
                    else:
                        sub_layout = item.layout()
                        if sub_layout is not None:
                            self.clear_layout(sub_layout)

    def refresh_dashboard_data(self):
        summary_data = self.fetch_generic_details("/dashboard-summary")
        if summary_data:
            self.dashboard_page.update_metrics(summary_data)
        low_stock_data = self.fetch_generic_details("/products/low-stock")
        self.dashboard_page.update_low_stock_alerts(low_stock_data, self.clear_layout)

    def refresh_clients_data(self):
        clients_data = self.fetch_generic_details("/clients")
        if not clients_data:
            self.clients_page.populate_table([])
            self.monthly_bills_page.update_client_dropdown([])
            return
        self.clients_list = clients_data
        self.clients_page.populate_table(clients_data)
        self.monthly_bills_page.update_client_dropdown(clients_data)
        
        try:
            # Clear search bar on refresh
            self.clients_page.search_bar.clear()
        except AttributeError:
            pass # Ignore if search bar not added yet

    def refresh_products_data(self):
        products_data = self.fetch_generic_details("/products")
        if not products_data:
            self.products_page.populate_table([])
            return
        self.products_list = products_data
        self.products_page.populate_table(products_data)
        # Reset search bar
        try:
            self.products_page.search_bar.clear()
        except AttributeError:
            pass # Ignore if search bar doesn't exist

    # ---
    # --- HELPER: Get Date Range from Filter UI ---
    # ---
    def _get_date_range_from_filter(self, page_key):
        """
        Reads the *stored* filter settings for a given page and returns (start_date, end_date) strings.
        Returns (None, None) if "All Time" is selected.
        """
        try:
            settings = self._filter_settings[page_key]
            filter_type = settings['type']

            today = datetime.today().date()

            if filter_type == "All Time":
                return None, None

            elif filter_type == "This Month":
                start_date = today.replace(day=1)
                next_month = (start_date + timedelta(days=32)).replace(day=1)
                end_date = next_month - timedelta(days=1)
                return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")

            elif filter_type == "This Week":
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)
                return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")

            elif filter_type == "Custom Date Range":
                start_date_str = settings['start'].toString("yyyy-MM-dd")
                end_date_str = settings['end'].toString("yyyy-MM-dd")
                return start_date_str, end_date_str

        except Exception as e:
            print(f"Warning: Could not read filter settings for {page_key}. {e}")
            return None, None

        return None, None

    # ---
    # --- refresh_orders_data ---
    # ---
    def refresh_orders_data(self, page_num=1):
        page_key = 'orders'
        self._current_page[page_key] = page_num

        base_endpoint = "/orders"
        
        # Use fetch_generic_details's params argument
        params = {}
        start_date_str, end_date_str = self._get_date_range_from_filter(page_key)
        if start_date_str:
            params["start_date"] = start_date_str
        if end_date_str:
            params["end_date"] = end_date_str

        params["page"] = self._current_page[page_key]
        params["per_page"] = 25

        response_data = self.fetch_generic_details(base_endpoint, params=params)

        if not response_data:
            self.orders_page.populate_table([])
            return

        orders_data = response_data.get('data', [])
        self._total_pages[page_key] = response_data.get('total_pages', 1)
        self._current_page[page_key] = response_data.get('current_page', 1)

        self.orders_page.populate_table(orders_data)

        try:
            # Clear search bar on refresh
            self.orders_page.search_bar.clear()
            page_text = f"Page {self._current_page[page_key]} of {self._total_pages[page_key]}"
            self.orders_page.page_label.setText(page_text)
            self.orders_page.prev_button.setEnabled(self._current_page[page_key] > 1)
            self.orders_page.next_button.setEnabled(self._current_page[page_key] < self._total_pages[page_key])
            # Update filter label
            self.orders_page.filter_label.setText(f"Filter: {self._filter_settings[page_key]['type']}")
        except AttributeError as e:
            print(f"Warning: Orders page pagination/filter UI not found. {e}")

    # ---
    # --- refresh_challans_data ---
    # ---
    def refresh_challans_data(self, page_num=1):
        page_key = 'challans'
        self._current_page[page_key] = page_num

        base_endpoint = "/challans"
        params = {}

        start_date_str, end_date_str = self._get_date_range_from_filter(page_key)
        if start_date_str:
            params["start_date"] = start_date_str
        if end_date_str:
            params["end_date"] = end_date_str

        params["page"] = self._current_page[page_key]
        params["per_page"] = 25

        response_data = self.fetch_generic_details(base_endpoint, params=params)

        if not response_data:
            self.challans_page.populate_table([])
            return

        challans_data = response_data.get('data', [])
        self._total_pages[page_key] = response_data.get('total_pages', 1)
        self._current_page[page_key] = response_data.get('current_page', 1)

        self.challans_page.populate_table(challans_data)

        try:
            # Clear search bar on refresh
            self.challans_page.search_bar.clear()
            page_text = f"Page {self._current_page[page_key]} of {self._total_pages[page_key]}"
            self.challans_page.page_label.setText(page_text)
            self.challans_page.prev_button.setEnabled(self._current_page[page_key] > 1)
            self.challans_page.next_button.setEnabled(self._current_page[page_key] < self._total_pages[page_key])
            # Update filter label
            self.challans_page.filter_label.setText(f"Filter: {self._filter_settings[page_key]['type']}")
        except AttributeError as e:
            print(f"Warning: Challans page pagination/filter UI not found. {e}")

    # ---
    # --- refresh_monthly_bills_data ---
    # ---
    def refresh_monthly_bills_data(self, page_num=1):
        page_key = 'monthly_bills'
        self._current_page[page_key] = page_num

        base_endpoint = "/monthly-bills"
        params = {}

        start_date_str, end_date_str = self._get_date_range_from_filter(page_key)
        if start_date_str:
            params["start_date"] = start_date_str
        if end_date_str:
            params["end_date"] = end_date_str

        params["page"] = self._current_page[page_key]
        params["per_page"] = 25

        response_data = self.fetch_generic_details(base_endpoint, params=params)

        if not response_data:
            self.monthly_bills_page.populate_table([])
            return

        bills_data = response_data.get('data', [])
        self._total_pages[page_key] = response_data.get('total_pages', 1)
        self._current_page[page_key] = response_data.get('current_page', 1)

        self.monthly_bills_page.populate_table(bills_data)

        try:
            # Clear search bar on refresh
            self.monthly_bills_page.search_bar.clear()
            page_text = f"Page {self._current_page[page_key]} of {self._total_pages[page_key]}"
            self.monthly_bills_page.page_label.setText(page_text)
            self.monthly_bills_page.prev_button.setEnabled(self._current_page[page_key] > 1)
            self.monthly_bills_page.next_button.setEnabled(self._current_page[page_key] < self._total_pages[page_key])
            # Update filter label
            self.monthly_bills_page.filter_label.setText(f"Filter: {self._filter_settings[page_key]['type']}")
        except AttributeError as e:
            print(f"Warning: Bills page pagination/filter UI not found. {e}")

    # ---
    # --- Generic Pagination and Filter Helpers ---
    # ---

    # NEW: Open the filter dialog
    def open_filter_dialog(self, page_key):
        """Opens the filter dialog and applies the new filter if accepted."""
        current_settings = self._filter_settings[page_key]
        dialog = FilterDialog(self, current_settings)

        if dialog.exec():
            # User clicked OK, save new settings
            self._filter_settings[page_key] = {
                'type': dialog.selected_filter_type,
                'start': dialog.start_date,
                'end': dialog.end_date
            }
            # Refresh the data for the current page
            refresh_method = getattr(self, f"refresh_{page_key}_data")
            refresh_method(page_num=1) # Reset to page 1 with new filter

    def go_to_next_page(self, page_key):
        """Advances to the next page for the given page_key."""
        if self._current_page[page_key] < self._total_pages[page_key]:
            refresh_method = getattr(self, f"refresh_{page_key}_data")
            refresh_method(page_num=self._current_page[page_key] + 1)

    def go_to_prev_page(self, page_key):
        """Goes to the previous page for the given page_key."""
        if self._current_page[page_key] > 1:
            refresh_method = getattr(self, f"refresh_{page_key}_data")
            refresh_method(page_num=self._current_page[page_key] - 1)

    def reset_page_filter(self, page_key):
        """Resets the filter UI and refreshes the data for the given page_key."""
        try:
            # UPDATED: Set default filter to "All Time"
            self._filter_settings[page_key] = {
                'type': 'All Time',
                'start': None,
                'end': None
            }
            # Update the label on the page
            page = getattr(self, f"{page_key}_page")
            page.filter_label.setText("Filter: All Time")

        except AttributeError as e:
            print(f"Warning: Could not reset filter UI for {page_key}. {e}")

        refresh_method = getattr(self, f"refresh_{page_key}_data")
        refresh_method(page_num=1)

    # --- End of new helpers ---

    def refresh_all_data(self):
        self.refresh_dashboard_data()
        self.refresh_products_data()
        self.refresh_clients_data()

        self.reset_page_filter('orders')
        self.reset_page_filter('challans')
        self.reset_page_filter('monthly_bills')

    def refresh_challans_and_orders(self):
        self.refresh_challans_data(page_num=self._current_page['challans'])
        self.refresh_orders_data(page_num=self._current_page['orders'])
        self.refresh_dashboard_data()

    # --- Action Methods (Products, Clients, Orders, Challans, Bills) ---
    
    # ===================================================================
    # --- UPDATED: view_product_details_by_id ---
    # ===================================================================
    def view_product_details_by_id(self, product_id):
        """
        Fetches product data and displays it in the new 
        ProductDetailDialog.
        """
        product_data = self.fetch_generic_details(f"/products/{product_id}")
        if not product_data: 
            QMessageBox.warning(self, "Error", "Could not fetch product details.")
            return
        
        # Create and execute the new dialog (now imported)
        dialog = ProductDetailDialog(product_data, self)
        dialog.exec()
    # ===================================================================

    def edit_product_by_id(self, product_id):
        product_data = self.fetch_generic_details(f"/products/{product_id}")
        if product_data:
            self.open_product_dialog(product_data)

    def delete_product_by_id(self, product_id):
        if self.confirm_delete("product", product_id):
            self.perform_delete(f"/products/{product_id}", "product", self.refresh_products_data)

    def open_product_dialog(self, product_data=None):
        dialog = ProductDialog(self, product_data)
        if dialog.exec():
            self.refresh_products_data()
            self.refresh_dashboard_data()

    def edit_client_by_id(self, client_id):
        client_data = self.fetch_generic_details(f"/clients/{client_id}")
        if client_data:
            self.open_client_dialog(client_data)

    def delete_client_by_id(self, client_id):
        if self.confirm_delete("client", client_id):
            self.perform_delete(f"/clients/{client_id}", "client", self.refresh_clients_data)

    def open_client_dialog(self, client_data=None):
        dialog = ClientDialog(self, client_data)
        if dialog.exec():
            self.refresh_clients_data()

    def open_client_pricing_window(self, client_id):
        client = self.fetch_generic_details(f"/clients/{client_id}")
        if not client: return
        dialog = ClientPricingDialog(self, client_id, client.get('company_name'))
        dialog.exec()

    def view_order_details(self, order_id):
        """Fetches order data and displays it in the new OrderDetailDialog."""
        order_data = self.fetch_generic_details(f"/orders/{order_id}")
        if not order_data:
            QMessageBox.warning(self, "Error", f"Could not fetch details for Order #{order_id}.")
            return

        # --- OLD QMessageBox Code (Removed) ---
        # msg = QMessageBox(self)
        # msg.setWindowTitle(f"Details for Order #{order_id}")
        # items_html = "<ul>"
        # for item in order_data.get('items', []):
        #     items_html += f"<li>{item['quantity']} x {item['product_name']} @ ₹{item['price_per_unit']:.2f}</li>"
        # items_html += "</ul>"
        # details_text = f"""
        # <b>Order ID:</b> {order_data.get('order_id')} <br>
        # <b>Client:</b> {order_data.get('client_name')} <br>
        # <b>Status:</b> {order_data.get('status')} <br>
        # <b>Total:</b> ₹{order_data.get('total_amount', 0):.2f} <br><br>
        # <b>Items:</b> {items_html}
        # """
        # msg.setText(details_text)
        # msg.exec()
        # --- End of Removed Code ---

        # --- NEW: Use OrderDetailDialog ---
        dialog = OrderDetailDialog(order_data, self)
        dialog.exec()
        # --- End of New Code ---

    def create_challan_for_order(self, order_id):
        if QMessageBox.question(self, "Confirm", f"Create a challan for Order ID: {order_id}?") == QMessageBox.StandardButton.Yes:
            try:
                response = requests.post(f"{API_BASE_URL}/challans", json={"order_id": order_id})
                response.raise_for_status()
                QMessageBox.information(self, "Success", response.json().get("message", "Challan created."))
                self.refresh_challans_and_orders()
            except requests.exceptions.RequestException as e:
                self.show_api_error("create challan", e)

    def delete_order_by_id(self, order_id):
         if self.confirm_delete("order", order_id):
            self.perform_delete(f"/orders/{order_id}", "order", lambda: (
                self.refresh_orders_data(page_num=self._current_page['orders']),
                self.refresh_dashboard_data(),
                self.refresh_products_data()
            ))

    def download_challan_pdf(self, challan_id):
         self.download_pdf(f"/challans/{challan_id}/pdf", f"Challan_{challan_id}.pdf")

    def reset_challan_billing(self, challan_id):
         if QMessageBox.question(self, "Confirm", f"Reset billing status for Challan ID: {challan_id}?") == QMessageBox.StandardButton.Yes:
            try:
                response = requests.post(f"{API_BASE_URL}/challans/{challan_id}/reset-billing")
                response.raise_for_status()
                QMessageBox.information(self, "Success", response.json().get("message", "Challan reset."))
                self.refresh_challans_data(page_num=self._current_page['challans'])
            except requests.exceptions.RequestException as e:
                self.show_api_error("reset challan", e)

    def delete_challan_by_id(self, challan_id):
         if self.confirm_delete("challan", challan_id):
            self.perform_delete(f"/challans/{challan_id}", "challan", self.refresh_challans_and_orders)

    def download_monthly_bill_pdf(self, bill_id):
        self.download_pdf(f"/monthly-bills/{bill_id}/pdf", f"Monthly_Bill_{bill_id}.pdf")

    def open_mark_as_paid_dialog(self, bill_id):
         dialog = MarkAsPaidDialog(self, bill_id)
         if dialog.exec():
            self.refresh_monthly_bills_data(page_num=self._current_page['monthly_bills'])

    def delete_monthly_bill_by_id(self, bill_id):
        if self.confirm_delete("monthly bill", bill_id):
            self.perform_delete(f"/monthly-bills/{bill_id}", "monthly bill", lambda: (
                self.refresh_monthly_bills_data(page_num=self._current_page['monthly_bills']),
                self.refresh_challans_data(page_num=self._current_page['challans']),
                self.refresh_dashboard_data()
            ))

    def generate_monthly_bill(self):
        client_id = self.monthly_bills_page.bill_client_combo.currentData()
        client_name = self.monthly_bills_page.bill_client_combo.currentText()
        month = self.monthly_bills_page.bill_month_combo.currentText()
        year_value = self.monthly_bills_page.bill_year_combo.value()
        if not client_id:
            QMessageBox.warning(self, "Input Error", "Please select a client.")
            return
        billing_month = f"{year_value}-{month}"
        
        if QMessageBox.question(self, "Confirm", f"Generate a bill for {client_name} for {billing_month}?") == QMessageBox.StandardButton.No:
            return
        try:
            payload = {"client_id": client_id, "billing_month": billing_month}
            response = requests.post(f"{API_BASE_URL}/monthly-bills", json=payload)
            response.raise_for_status()
            # Check for 'no action' message vs 'success' message
            json_response = response.json()
            if "bill_id" in json_response:
                QMessageBox.information(self, "Success", json_response.get("message", "Bill generation complete."))
            else:
                QMessageBox.information(self, "Info", json_response.get("message", "No new challans to bill."))

            self.refresh_monthly_bills_data(page_num=1)
            self.refresh_challans_data(page_num=1)
            self.refresh_dashboard_data()
            
            # --- NEW: Re-check the bill status after generation ---
            self.monthly_bills_page.run_bill_check()
            
        except requests.exceptions.RequestException as e:
            self.show_api_error("generate bill", e)

    # ---
    # --- CSV/XLSX Export Functions ---
    # ---
    def export_table_to_csv(self, table_widget, default_filename):
        if table_widget.rowCount() == 0:
            QMessageBox.information(self, "No Data", "There is no data to export.")
            return
        save_path, _ = QFileDialog.getSaveFileName(self, "Save CSV", default_filename, "CSV Files (*.csv)")
        if not save_path:
            return
        try:
            with open(save_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                headers = []
                col_indices = []
                for col in range(table_widget.columnCount()):
                    header_text = table_widget.horizontalHeaderItem(col).text()
                    if header_text != "Actions" and header_text != "":
                        headers.append(header_text)
                        col_indices.append(col)
                writer.writerow(headers)
                for row in range(table_widget.rowCount()):
                    if table_widget.isRowHidden(row):
                        continue
                    row_data = []
                    for col in col_indices:
                        item = table_widget.item(row, col)
                        row_data.append(item.text() if item else "")
                    writer.writerow(row_data)
            QMessageBox.information(self, "Success", f"Data exported successfully to:\n{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export data: {e}")

    def export_table_to_xlsx(self, table_widget, default_filename):
        if table_widget.rowCount() == 0:
            QMessageBox.information(self, "No Data", "There is no data to export.")
            return
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", default_filename, "Excel Files (*.xlsx)")
        if not save_path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = []
            col_indices = []
            for col in range(table_widget.columnCount()):
                header_text = table_widget.horizontalHeaderItem(col).text()
                if header_text != "Actions" and header_text != "":
                    headers.append(header_text)
                    col_indices.append(col)
            ws.append(headers)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = openpyxl.styles.Font(bold=True)
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(len(header) + 2, 15)
            for row in range(table_widget.rowCount()):
                if table_widget.isRowHidden(row):
                    continue
                row_data = []
                for col_idx, col in enumerate(col_indices, 1):
                    item = table_widget.item(row, col)
                    cell_value = item.text() if item else ""
                    if isinstance(cell_value, str):
                        cell_value = cell_value.replace('₹', '').replace(',', '')
                    try:
                        numeric_value = int(cell_value)
                        row_data.append(numeric_value)
                    except ValueError:
                        try:
                            numeric_value = float(cell_value)
                            row_data.append(numeric_value)
                        except ValueError:
                            row_data.append(item.text() if item else "")
                ws.append(row_data)
            wb.save(save_path)
            QMessageBox.information(self, "Success", f"Data exported successfully to:\n{save_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export data: {e}")

    # --- Generic Helper Functions ---
    def confirm_delete(self, item_name, item_id):
        return QMessageBox.warning(self, "Confirm Delete",
                                   f"Are you sure you want to permanently delete this {item_name} (ID: {item_id})? This cannot be undone.",
                                   QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes

    def perform_delete(self, endpoint, item_name, refresh_callback):
         try:
            response = requests.delete(f"{API_BASE_URL}{endpoint}")
            response.raise_for_status()
            QMessageBox.information(self, "Success", response.json().get("message", f"{item_name.title()} deleted."))
            refresh_callback()
         except requests.exceptions.RequestException as e:
            self.show_api_error(f"delete {item_name}", e)

    def show_api_error(self, action, e):
         try:
            error = e.response.json().get('error') if e.response and e.response.content else str(e)
         except:
            error = str(e)
         QMessageBox.critical(self, "API Error", f"Failed to {action}: {error}")

    def download_pdf(self, endpoint, default_filename):
         try:
            response = requests.get(f"{API_BASE_URL}{endpoint}")
            response.raise_for_status()
            save_path, _ = QFileDialog.getSaveFileName(self, "Save PDF", default_filename, "PDF Files (*.pdf)")
            if save_path:
                with open(save_path, 'wb') as f:
                    f.write(response.content)
                QMessageBox.information(self, "Success", f"PDF saved to:\n{save_path}")
         except requests.exceptions.RequestException as e:
            self.show_api_error("download PDF", e)
         except Exception as e:
             QMessageBox.critical(self, "File Error", f"Could not save or open PDF file: {e}")

    # ---
    # --- STYLESHEET (UPDATED) ---
    # ---
    def apply_styles(self):
            COLOR_BG_SIDEBAR = "#483D8B"
            COLOR_BG_CONTENT = "#F8F8FF"
            COLOR_BG_CARD = "#FFFFFF"
            COLOR_BG_TABLE_ALT = "#E6E6FA"
            COLOR_TEXT_LIGHT = "#FFFFFF"
            COLOR_TEXT_DARK = "#333333"
            COLOR_TEXT_SIDEBAR_INACTIVE = "#D8BFD8"
            COLOR_ACCENT_GOLD = "#DAA520"
            COLOR_ACCENT_BLUE_ACTIVE = "#6A5ACD"
            COLOR_BORDER = "#DCDCDC"
            self.setStyleSheet(f"""
                QMainWindow {{ background-color: {COLOR_BG_CONTENT}; }}
                #SidebarFrame {{ background-color: {COLOR_BG_SIDEBAR}; }}
                #ContentFrame {{ 
                    background-color: {COLOR_BG_CONTENT}; 
                    padding: 20px; /* <--- ADDED PADDING HERE */ 
                }}
                #BrandLabel {{ color: {COLOR_ACCENT_GOLD}; font-size: 18pt; font-weight: bold; padding: 10px; margin-bottom: 10px; }}
                #SidebarButton {{ background-color: transparent; color: {COLOR_TEXT_SIDEBAR_INACTIVE}; border: none; padding: 12px; font-size: 11pt; text-align: left; border-radius: 5px; }}
                #SidebarButton:hover {{ background-color: {COLOR_ACCENT_BLUE_ACTIVE}; color: {COLOR_TEXT_LIGHT}; }}
                #SidebarButton:checked {{ background-color: {COLOR_ACCENT_BLUE_ACTIVE}; color: {COLOR_TEXT_LIGHT}; font-weight: bold; }}
                QWidget {{ background-color: {COLOR_BG_CONTENT}; color: {COLOR_TEXT_DARK}; font-family: Helvetica, Arial, sans-serif; font-size: 10pt; }}
                #ContentFrame > QWidget {{ background-color: {COLOR_BG_CONTENT}; }}
                QLabel#Header {{ font-size: 18pt; font-weight: bold; padding: 10px 0; color: {COLOR_BG_SIDEBAR}; margin-bottom: 10px; background-color: transparent; }}
                QLabel#MetricNumber {{ font-size: 24pt; font-weight: bold; color: {COLOR_TEXT_DARK}; background-color: transparent; }}
                QLabel {{ background-color: transparent; color: {COLOR_TEXT_DARK}; }}

                /* 1. Default CardFrame style (for Generate Bill, Dashboard Metrics, etc.) */
                QFrame[objectName="CardFrame"] {{
                    background-color: {COLOR_BG_CARD}; /* Ensure distinct background */
                    border: 1px solid {COLOR_BORDER};
                    border-radius: 8px;  /* <--- Rounded corners */
                    padding: 15px;     /* Existing padding */
                    margin-bottom: 15px; /* Existing margin */
                }}
                /* Make labels inside default cards have transparent BG (relative to card) */
                QFrame[objectName="CardFrame"] QLabel {{
                    background-color: transparent; /* Changed from COLOR_BG_CARD */
                    border: none; /* Ensure no extra borders on labels inside */
                    padding: 0;   /* Ensure no extra padding on labels inside */
                }}

                /* 2. Style for the smaller Filter Card (used on Bill page) */
                QFrame[objectName="FilterCardFrame"] {{
                    background-color: {COLOR_BG_CARD};
                    border: 1px solid {COLOR_BORDER};
                    border-radius: 5px;
                    padding: 5px 10px;
                    margin-bottom: 15px;
                }}
                /* Make labels *inside the filter card* flat */
                QFrame[objectName="FilterCardFrame"] QLabel {{
                    background-color: {COLOR_BG_CARD};
                    border: none;
                    padding: 0;
                }}


                /* Style for main action buttons (e.g., old Generate Bill text button) */
                QPushButton {{ background-color: {COLOR_BG_SIDEBAR}; color: {COLOR_TEXT_LIGHT}; padding: 5px 10px; border: 1px solid {COLOR_BG_SIDEBAR}; border-radius: 4px; font-weight: bold; min-height: 20px; }}
                QPushButton:hover {{ background-color: {COLOR_ACCENT_BLUE_ACTIVE}; border: 1px solid {COLOR_ACCENT_BLUE_ACTIVE}; }}
                QPushButton:disabled {{ background-color: #CCCCCC; border: 1px solid #BBBBBB; color: #888888; }}

                /* Style for filter icon buttons (funnel, reset) */
                QPushButton[objectName="FilterButton"] {{
                    background-color: {COLOR_BG_TABLE_ALT};
                    color: {COLOR_BG_SIDEBAR};
                    border: 1px solid {COLOR_BORDER};
                    min-width: 30px;
                    max-width: 30px;
                    min-height: 28px;
                    max-height: 28px;
                    padding: 0px;
                }}
                QPushButton[objectName="FilterButton"]:hover {{
                    background-color: {COLOR_BG_SIDEBAR};
                    color: {COLOR_TEXT_LIGHT};
                    border: 1px solid {COLOR_BG_SIDEBAR};
                }}

                /* Style for Export icon buttons (CSV, Excel) */
                QPushButton[toolTip="Export to CSV"], QPushButton[toolTip="Export to Excel"] {{
                    background-color: transparent;
                    border: 1px solid {COLOR_BORDER};
                    color: {COLOR_BG_SIDEBAR};
                    min-width: 40px;
                    max-width: 40px;
                    min-height: 40px;
                    max-height: 40px;
                    padding: 0px;
                }}
                QPushButton[toolTip="Export to CSV"]:hover, QPushButton[toolTip="Export to Excel"]:hover {{
                    background-color: {COLOR_BG_TABLE_ALT};
                    border: 1px solid {COLOR_BG_SIDEBAR};
                }}

                /* --- NEW STYLE: Add icon buttons --- */
                QPushButton[objectName="AddButton"] {{
                    background-color: transparent;
                    border: 1px solid {COLOR_BORDER};
                    color: {COLOR_BG_SIDEBAR}; /* Icon color */
                    min-width: 40px;
                    max-width: 40px;
                    min-height: 40px;
                    max-height: 40px;
                    padding: 0px;
                    border-radius: 4px; /* Optional: make it slightly rounded */
                }}
                QPushButton[objectName="AddButton"]:hover {{
                    background-color: {COLOR_BG_TABLE_ALT};
                    border: 1px solid {COLOR_BG_SIDEBAR};
                }}
                /* --- END NEW STYLE --- */

                QTableWidget {{ background-color: {COLOR_BG_CARD}; border: 1px solid {COLOR_BORDER}; alternate-background-color: {COLOR_BG_TABLE_ALT}; gridline-color: {COLOR_BORDER}; selection-background-color: {COLOR_ACCENT_GOLD}; selection-color: {COLOR_TEXT_DARK}; color: {COLOR_TEXT_DARK}; }}
                QHeaderView::section {{ background-color: {COLOR_BG_TABLE_ALT}; padding: 8px; border-top: none; border-bottom: 2px solid {COLOR_BG_SIDEBAR}; border-left: none; border-right: 1px solid {COLOR_BORDER}; font-weight: bold; color: {COLOR_BG_SIDEBAR}; }}
                
                /* Make header clickable */
                QHeaderView::section:hover {{ background-color: #D8BFD8; }}


                /* Style for table icon buttons (View, Delete) */
                QTableWidget QPushButton {{
                    font-size: 11pt; padding: 2px 4px; border: none; background-color: transparent;
                    font-weight: normal; color: {COLOR_ACCENT_BLUE_ACTIVE};
                    min-width: 25px; max-width: 25px; min-height: 20px;
                }}
                QTableWidget QPushButton:hover {{
                    background-color: {COLOR_BG_TABLE_ALT}; color: {COLOR_BG_SIDEBAR}; border-radius: 3px;
                }}
                QTableWidget QPushButton:disabled {{ color: #AAAAAA; background-color: transparent; }}

                /* Style for table text buttons (e.g. Mark as Paid) */
                QTableWidget QPushButton[toolTip="Mark as Paid"] {{
                    background-color: {COLOR_BG_SIDEBAR};
                    color: {COLOR_TEXT_LIGHT};
                    border: none;
                    border-radius: 4px;
                    padding: 4px 8px;
                    font-size: 9pt;
                    font-weight: bold;
                    min-width: 25px;
                    max-width: 25px;
                }}
                QTableWidget QPushButton[toolTip="Mark as Paid"]:hover {{
                    background-color: {COLOR_ACCENT_BLUE_ACTIVE};
                }}
                QTableWidget QPushButton[toolTip="Mark as Paid"]:disabled {{
                    background-color: #CCCCCC;
                    color: #888888;
                }}

                /* This rule now correctly styles the Generate Bill widgets AND search bars */
                QComboBox, QLineEdit, QDateEdit, QSpinBox, QDoubleSpinBox, QTextEdit {{
                    background-color: {COLOR_BG_CARD};
                    border: 1px solid {COLOR_BORDER};
                    padding: 5px;
                    border-radius: 4px;
                    min-height: 20px;
                    color: {COLOR_TEXT_DARK};
                }}
                
                /* Style for search bar */
                QLineEdit[objectName="SearchBar"] {{
                    background-color: {COLOR_BG_CARD};
                    padding: 5px 10px;
                    border-radius: 15px; /* Make it rounded */
                    min-height: 22px;
                    border: 1px solid {COLOR_BORDER};
                }}
                QLineEdit[objectName="SearchBar"]:focus {{
                    border: 1px solid {COLOR_BG_SIDEBAR};
                }}

                QComboBox::drop-down {{ border: none; }}
                QComboBox::down-arrow {{ width: 10px; }}
                QDateEdit {{ padding-right: 15px; }}
                
                /* ---
                --- BaseDialog Custom Styles (NEW) ---
                --- */
                
                /* The main background frame of the dialog */
                #DialogBackground {{
                    background-color: {COLOR_BG_CONTENT};
                    border: 1px solid {COLOR_BORDER};
                    border-radius: 8px;
                }}

                /* The custom header/title bar */
                #DialogHeader {{
                    background-color: {COLOR_BG_SIDEBAR};
                    /* Only round the top corners */
                    border-top-left-radius: 7px;
                    border-top-right-radius: 7px;
                    border-bottom: 1px solid {COLOR_BG_SIDEBAR};
                }}

                /* The title text in the header */
                #DialogHeaderTitle {{
                    color: {COLOR_TEXT_LIGHT};
                    font-size: 11pt;
                    font-weight: bold;
                    background-color: transparent;
                }}
                
                /* The icon in the header */
                #DialogHeader QLabel {{
                    background-color: transparent;
                }}

                /* The main content area */
                #DialogContent {{
                    background-color: {COLOR_BG_CONTENT};
                }}
                
                /* All widgets inside the content area */
                #DialogContent QLabel, 
                #DialogContent QLineEdit, 
                #DialogContent QTextEdit, 
                #DialogContent QSpinBox, 
                #DialogContent QDoubleSpinBox,
                #DialogContent QDateEdit,
                #DialogContent QComboBox {{
                    font-size: 10pt; 
                    color: {COLOR_TEXT_DARK}; 
                    background-color: transparent;
                }}

                /* --- STYLES FOR DETAIL DIALOG (UPDATED) --- */
                
                /* NEW: Frame for the product details form */
                #DialogContent QFrame[objectName="DetailFormFrame"] {{
                    background-color: #FFFFFF;
                    border: 1px solid {COLOR_BORDER};
                    border-radius: 5px;
                    margin-top: 10px; /* Space from image */
                }}
                
                /* All labels inside the detail frame should be on white bg */
                #DialogContent QFrame[objectName="DetailFormFrame"] QLabel {{
                    background-color: #FFFFFF;
                }}

                /* NEW: Style for "Key" labels in detail dialog (e.g., "Name:") */
                #DialogContent QLabel[objectName="DetailKey"] {{
                    font-weight: bold;
                    color: #555555; /* Slightly muted color */
                }}

                /* NEW: Style for "Value" labels in detail dialog (e.g., "Harpic Power") */
                #DialogContent QLabel[objectName="DetailValue"] {{
                    font-weight: normal;
                    padding: 2px;
                    color: {COLOR_TEXT_DARK}; /* Ensure it's the main dark color */
                }}

                /* --- END DETAIL DIALOG STYLES --- */

                /* The custom footer/button bar area */
                #DialogFooter {{
                    background-color: #F1F1F1; /* A slightly different BG for separation */
                    border-top: 1px solid {COLOR_BORDER};
                    /* Only round the bottom corners */
                    border-bottom-left-radius: 7px;
                    border-bottom-right-radius: 7px;
                }}

                /* Style the buttons inside our custom footer */
                #DialogFooter QPushButton {{
                    background-color: {COLOR_BG_SIDEBAR}; 
                    color: {COLOR_TEXT_LIGHT};
                    padding: 5px 15px; /* Make buttons a bit wider */
                    border-radius: 4px;
                    font-weight: bold;
                    min-height: 20px;
                }}
                #DialogFooter QPushButton:hover {{
                    background-color: {COLOR_ACCENT_BLUE_ACTIVE};
                }}

                /* --- End BaseDialog Styles --- */
                
                QMessageBox {{ font-size: 10pt; background-color: {COLOR_BG_CONTENT}; }}
                QMessageBox QLabel {{ font-size: 10pt; color: {COLOR_TEXT_DARK}; background-color: transparent; }}
            """)
            
    # --- Reusable Table Style Helper ---
    def setup_table_style(self, table):
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        table.verticalHeader().setVisible(False)
        table.setAlternatingRowColors(True)
        
        # --- NEW: Enable Sorting ---
        table.setSortingEnabled(True)
        # Set default sort arrow to be down
        table.horizontalHeader().setSortIndicatorShown(True)
        table.horizontalHeader().setSortIndicator(0, Qt.SortOrder.AscendingOrder)


# --- Main execution block ---
if __name__ == '__main__':
    app = QApplication(sys.argv)

    if not os.path.isdir(ICON_PATH):
         print(f"Warning: Icon folder '{ICON_PATH}' not found. Icons will not be displayed.")

    window = AdminDashboard()
    window.showMaximized()
    sys.exit(app.exec())